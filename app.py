# -*- coding: utf-8 -*-
"""
异常信息填报系统 - Flask主程序
沙钢永兴特钢技术质量处验收班使用
支持 PostgreSQL (生产) 和 SQLite (本地开发)
"""

import os
import sqlite3
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, session, make_response, send_from_directory
from functools import wraps

app = Flask(__name__)

# 生产环境配置
DEBUG = os.environ.get('DEBUG', 'False').lower() == 'true'
SECRET_KEY = os.environ.get('SECRET_KEY', 'shagang_yongxing_anomaly_system_2024_fallback')
app.secret_key = SECRET_KEY

# 数据库配置
DATABASE_URL = os.environ.get('DATABASE_URL')
SQLITE_DB = 'data.db'

# 常用物料名称列表
MATERIAL_LIST = [
    '钒铁', '钼铁', '铝块', '铝粒', '钒氮合金', '钛铁', '铌铁', '硼铁'
]

# 初始用户数据
INITIAL_USERS = [
    {'name': '郭鑫', 'role': 'admin', 'avatar': '👨‍💼'},
    {'name': '赵飞', 'role': 'statistician', 'avatar': '📊'},
    {'name': '郭娟娟', 'role': 'inspector', 'avatar': '👩‍🔧'},
    {'name': '姬银平', 'role': 'inspector', 'avatar': '👩‍🔧'},
    {'name': '王三妮', 'role': 'inspector', 'avatar': '👩‍🔧'},
]


def get_db_type():
    """判断数据库类型"""
    return 'postgresql' if DATABASE_URL else 'sqlite'


def get_db():
    """获取数据库连接（兼容 SQLite 和 PostgreSQL）"""
    if DATABASE_URL:
        import psycopg2
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = True
        return conn
    else:
        conn = sqlite3.connect(SQLITE_DB)
        conn.row_factory = sqlite3.Row
        return conn


def init_db():
    """初始化数据库 - 自动检测并创建表"""
    db_type = get_db_type()
    
    if db_type == 'postgresql':
        import psycopg2
        from psycopg2.extras import RealDictCursor
        
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 检查表是否存在
        cursor.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_name = 'users'
            )
        """)
        table_exists = cursor.fetchone()['exists']
        
        if not table_exists:
            # 创建用户表
            cursor.execute('''
                CREATE TABLE users (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(100) UNIQUE NOT NULL,
                    role VARCHAR(20) NOT NULL,
                    avatar VARCHAR(50) DEFAULT '',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # 创建异常记录表
            cursor.execute('''
                CREATE TABLE records (
                    id SERIAL PRIMARY KEY,
                    date VARCHAR(10) NOT NULL,
                    material_name VARCHAR(100),
                    supplier VARCHAR(200),
                    truck_no VARCHAR(50),
                    abnormal_desc TEXT,
                    inspector1 VARCHAR(100),
                    inspector2 VARCHAR(100),
                    created_by VARCHAR(100),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # 创建索引
            cursor.execute('CREATE INDEX idx_records_date ON records(date)')
            cursor.execute('CREATE INDEX idx_records_inspector1 ON records(inspector1)')
            cursor.execute('CREATE INDEX idx_records_inspector2 ON records(inspector2)')
            
            conn.commit()
        
        # 插入初始用户
        for user in INITIAL_USERS:
            try:
                cursor.execute(
                    'INSERT INTO users (name, role, avatar) VALUES (%s, %s, %s) ON CONFLICT (name) DO NOTHING',
                    (user['name'], user['role'], user['avatar'])
                )
            except Exception as e:
                pass
        
        conn.commit()
        cursor.close()
        conn.close()
        
    else:
        # SQLite 初始化
        conn = sqlite3.connect(SQLITE_DB)
        cursor = conn.cursor()
        
        # 创建用户表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                role TEXT NOT NULL,
                avatar TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 创建异常记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                material_name TEXT,
                supplier TEXT,
                truck_no TEXT,
                abnormal_desc TEXT,
                inspector1 TEXT,
                inspector2 TEXT,
                created_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # 插入初始用户
        for user in INITIAL_USERS:
            try:
                cursor.execute(
                    'INSERT OR IGNORE INTO users (name, role, avatar) VALUES (?, ?, ?)',
                    (user['name'], user['role'], user['avatar'])
                )
            except:
                pass
        
        conn.commit()
        conn.close()


def login_required(f):
    """登录验证装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return jsonify({'error': '请先登录'}), 401
        return f(*args, **kwargs)
    return decorated_function


def get_user_list():
    """获取所有用户列表"""
    db_type = get_db_type()
    conn = get_db()
    
    if db_type == 'postgresql':
        from psycopg2.extras import RealDictCursor
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute('SELECT id, name, role, COALESCE(avatar, \'\') as avatar FROM users ORDER BY id')
        users = [dict(row) for row in cursor.fetchall()]
        cursor.close()
    else:
        cursor = conn.cursor()
        cursor.execute('SELECT id, name, role, COALESCE(avatar, \'\') as avatar FROM users ORDER BY id')
        users = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return users


def get_records_by_user(username, role):
    """根据用户角色获取记录"""
    db_type = get_db_type()
    conn = get_db()
    
    if db_type == 'postgresql':
        from psycopg2.extras import RealDictCursor
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        if role == 'admin' or role == 'statistician':
            cursor.execute('SELECT * FROM records ORDER BY date DESC, id DESC')
        else:
            cursor.execute('''
                SELECT * FROM records 
                WHERE inspector1 = %s OR inspector2 = %s
                ORDER BY date DESC, id DESC
            ''', (username, username))
        
        records = [dict(row) for row in cursor.fetchall()]
        cursor.close()
    else:
        cursor = conn.cursor()
        
        if role == 'admin' or role == 'statistician':
            cursor.execute('SELECT * FROM records ORDER BY date DESC, id DESC')
        else:
            cursor.execute('''
                SELECT * FROM records 
                WHERE inspector1 = ? OR inspector2 = ?
                ORDER BY date DESC, id DESC
            ''', (username, username))
        
        records = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return records


def add_record(data):
    """添加新记录"""
    db_type = get_db_type()
    
    if db_type == 'postgresql':
        from psycopg2.extras import RealDictCursor
        import psycopg2
        
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute('''
            INSERT INTO records (date, material_name, supplier, truck_no, abnormal_desc, inspector1, inspector2, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        ''', (
            data['date'],
            data['material_name'],
            data['supplier'],
            data['truck_no'],
            data['abnormal_desc'],
            data['inspector1'],
            data['inspector2'],
            data['created_by']
        ))
        
        result = cursor.fetchone()
        record_id = result['id']
        conn.commit()
        cursor.close()
        conn.close()
    else:
        conn = sqlite3.connect(SQLITE_DB)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO records (date, material_name, supplier, truck_no, abnormal_desc, inspector1, inspector2, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['date'],
            data['material_name'],
            data['supplier'],
            data['truck_no'],
            data['abnormal_desc'],
            data['inspector1'],
            data['inspector2'],
            data['created_by']
        ))
        
        record_id = cursor.lastrowid
        conn.commit()
        conn.close()
    
    return record_id


def update_record(record_id, data):
    """更新记录"""
    db_type = get_db_type()
    
    if db_type == 'postgresql':
        import psycopg2
        from psycopg2.extras import RealDictCursor
        
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute('''
            UPDATE records 
            SET date = %s, material_name = %s, supplier = %s, truck_no = %s, 
                abnormal_desc = %s, inspector1 = %s, inspector2 = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        ''', (
            data['date'],
            data['material_name'],
            data['supplier'],
            data['truck_no'],
            data['abnormal_desc'],
            data['inspector1'],
            data['inspector2'],
            record_id
        ))
        
        conn.commit()
        rowcount = cursor.rowcount
        cursor.close()
        conn.close()
    else:
        conn = sqlite3.connect(SQLITE_DB)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE records 
            SET date = ?, material_name = ?, supplier = ?, truck_no = ?, 
                abnormal_desc = ?, inspector1 = ?, inspector2 = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (
            data['date'],
            data['material_name'],
            data['supplier'],
            data['truck_no'],
            data['abnormal_desc'],
            data['inspector1'],
            data['inspector2'],
            record_id
        ))
        
        conn.commit()
        rowcount = cursor.rowcount
        conn.close()
    
    return rowcount > 0


def delete_record(record_id):
    """删除记录"""
    db_type = get_db_type()
    
    if db_type == 'postgresql':
        import psycopg2
        
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM records WHERE id = %s', (record_id,))
        conn.commit()
        rowcount = cursor.rowcount
        cursor.close()
        conn.close()
    else:
        conn = sqlite3.connect(SQLITE_DB)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM records WHERE id = ?', (record_id,))
        conn.commit()
        rowcount = cursor.rowcount
        conn.close()
    
    return rowcount > 0


def get_record_by_id(record_id):
    """根据ID获取单条记录"""
    db_type = get_db_type()
    conn = get_db()
    
    if db_type == 'postgresql':
        from psycopg2.extras import RealDictCursor
        import psycopg2
        
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        cursor.execute('SELECT * FROM records WHERE id = %s', (record_id,))
        record = cursor.fetchone()
        record = dict(record) if record else None
        cursor.close()
        conn.close()
    else:
        conn = sqlite3.connect(SQLITE_DB)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM records WHERE id = ?', (record_id,))
        record = cursor.fetchone()
        record = dict(record) if record else None
        conn.close()
    
    return record


def export_to_excel(records, filename):
    """导出记录到Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "异常信息记录"
    
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    
    # 边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ['序号', '日期', '物料名称', '供货单位', '车号', '异常情况', '验收员1', '验收员2', '填报人', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
    
    # 写入数据
    for row_num, record in enumerate(records, 2):
        ws.cell(row=row_num, column=1, value=row_num-1).border = thin_border
        ws.cell(row=row_num, column=2, value=record.get('date', '')).border = thin_border
        ws.cell(row=row_num, column=3, value=record.get('material_name', '')).border = thin_border
        ws.cell(row=row_num, column=4, value=record.get('supplier', '')).border = thin_border
        ws.cell(row=row_num, column=5, value=record.get('truck_no', '')).border = thin_border
        ws.cell(row=row_num, column=6, value=record.get('abnormal_desc', '')).border = thin_border
        ws.cell(row=row_num, column=7, value=record.get('inspector1', '')).border = thin_border
        ws.cell(row=row_num, column=8, value=record.get('inspector2', '')).border = thin_border
        ws.cell(row=row_num, column=9, value=record.get('created_by', '')).border = thin_border
        ws.cell(row=row_num, column=10, value=str(record.get('created_at', ''))[:19] if record.get('created_at') else '').border = thin_border
    
    # 设置列宽
    widths = [8, 12, 12, 20, 10, 40, 10, 10, 10, 20]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = width
    
    wb.save(filename)
    return filename


def get_dashboard_stats():
    """获取仪表盘统计数据"""
    db_type = get_db_type()
    conn = get_db()
    
    today = datetime.now().strftime('%Y-%m-%d')
    week_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    month_start = datetime.now().strftime('%Y-%m') + '-01'
    week_ago_prev = (datetime.now() - timedelta(days=14)).strftime('%Y-%m-%d')
    
    stats = {
        'week_count': 0,
        'week_change': 0,
        'month_count': 0,
        'pending_count': 0,
        'material_types': 0,
        'trend_data': [],
        'material_dist': []
    }
    
    if db_type == 'postgresql':
        from psycopg2.extras import RealDictCursor
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 本周异常数
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE date >= %s', (week_ago,))
        stats['week_count'] = cursor.fetchone()['count']
        
        # 上周异常数（用于计算趋势）
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE date >= %s AND date < %s', (week_ago_prev, week_ago))
        prev_week = cursor.fetchone()['count']
        stats['week_change'] = stats['week_count'] - prev_week
        
        # 本月异常数
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE date >= %s', (month_start,))
        stats['month_count'] = cursor.fetchone()['count']
        
        # 涉及物料种类
        cursor.execute('SELECT COUNT(DISTINCT material_name) as count FROM records WHERE material_name IS NOT NULL AND material_name != \'\'')
        stats['material_types'] = cursor.fetchone()['count']
        
        # 最近7天趋势
        for i in range(6, -1, -1):
            day = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
            cursor.execute('SELECT COUNT(*) as count FROM records WHERE date = %s', (day,))
            count = cursor.fetchone()['count']
            stats['trend_data'].append({
                'date': day[5:],  # MM-DD格式
                'count': count
            })
        
        # 物料分布
        cursor.execute('''
            SELECT material_name, COUNT(*) as count 
            FROM records 
            WHERE material_name IS NOT NULL AND material_name != ''
            GROUP BY material_name 
            ORDER BY count DESC
        ''')
        for row in cursor.fetchall():
            stats['material_dist'].append({
                'name': row['material_name'],
                'value': row['count']
            })
        
        cursor.close()
    else:
        cursor = conn.cursor()
        
        # 本周异常数
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE date >= ?', (week_ago,))
        stats['week_count'] = cursor.fetchone()['count']
        
        # 上周异常数
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE date >= ? AND date < ?', (week_ago_prev, week_ago))
        prev_week = cursor.fetchone()['count']
        stats['week_change'] = stats['week_count'] - prev_week
        
        # 本月异常数
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE date >= ?', (month_start,))
        stats['month_count'] = cursor.fetchone()['count']
        
        # 涉及物料种类
        cursor.execute('SELECT COUNT(DISTINCT material_name) as count FROM records WHERE material_name IS NOT NULL AND material_name != \'\'')
        stats['material_types'] = cursor.fetchone()['count']
        
        # 最近7天趋势
        for i in range(6, -1, -1):
            day = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
            cursor.execute('SELECT COUNT(*) as count FROM records WHERE date = ?', (day,))
            count = cursor.fetchone()['count']
            stats['trend_data'].append({
                'date': day[5:],
                'count': count
            })
        
        # 物料分布
        cursor.execute('''
            SELECT material_name, COUNT(*) as count 
            FROM records 
            WHERE material_name IS NOT NULL AND material_name != ''
            GROUP BY material_name 
            ORDER BY count DESC
        ''')
        for row in cursor.fetchall():
            stats['material_dist'].append({
                'name': row['material_name'],
                'value': row['count']
            })
    
    conn.close()
    return stats


def get_user_stats(username):
    """获取用户个人统计"""
    db_type = get_db_type()
    conn = get_db()
    
    month_start = datetime.now().strftime('%Y-%m') + '-01'
    
    stats = {'month_count': 0, 'total_count': 0}
    
    if db_type == 'postgresql':
        from psycopg2.extras import RealDictCursor
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE created_by = %s AND date >= %s', (username, month_start))
        stats['month_count'] = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE created_by = %s', (username,))
        stats['total_count'] = cursor.fetchone()['count']
        
        cursor.close()
    else:
        cursor = conn.cursor()
        
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE created_by = ? AND date >= ?', (username, month_start))
        stats['month_count'] = cursor.fetchone()['count']
        
        cursor.execute('SELECT COUNT(*) as count FROM records WHERE created_by = ?', (username,))
        stats['total_count'] = cursor.fetchone()['count']
    
    conn.close()
    return stats


# ==================== 路由 ====================

@app.route('/')
def index():
    """主页 - SPA应用"""
    return render_template('index.html')


@app.route('/manifest.json')
def manifest():
    """PWA清单"""
    return send_from_directory('.', 'manifest.json')


@app.route('/sw.js')
def service_worker():
    """Service Worker"""
    return send_from_directory('.', 'sw.js')


# ==================== API接口 ====================

@app.route('/api/materials')
def api_materials():
    """获取物料列表"""
    return jsonify({'materials': MATERIAL_LIST})


@app.route('/api/users')
def api_users():
    """获取用户列表"""
    users = get_user_list()
    return jsonify({'users': users})


@app.route('/api/login', methods=['POST'])
def api_login():
    """登录"""
    data = request.get_json()
    username = data.get('name')
    
    if not username:
        return jsonify({'error': '请选择用户'}), 400
    
    users = get_user_list()
    user = next((u for u in users if u['name'] == username), None)
    
    if not user:
        return jsonify({'error': '用户不存在'}), 404
    
    session['user'] = {
        'id': user['id'],
        'name': user['name'],
        'role': user['role'],
        'avatar': user.get('avatar', '')
    }
    
    return jsonify({'success': True, 'user': session['user']})


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """退出登录"""
    session.pop('user', None)
    return jsonify({'success': True})


@app.route('/api/me')
def api_me():
    """获取当前用户信息"""
    if 'user' not in session:
        return jsonify({'error': '未登录'}), 401
    return jsonify({'user': session['user']})


@app.route('/api/records')
@login_required
def api_records():
    """获取记录列表"""
    user = session['user']
    records = get_records_by_user(user['name'], user['role'])
    return jsonify({'records': records})


@app.route('/api/records', methods=['POST'])
@login_required
def api_create_record():
    """创建记录"""
    data = request.get_json()
    user = session['user']
    
    # 验证必填字段
    required = ['date', 'material_name', 'supplier', 'truck_no', 'abnormal_desc', 'inspector1', 'inspector2']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'缺少必填字段: {field}'}), 400
    
    data['created_by'] = user['name']
    
    try:
        record_id = add_record(data)
        return jsonify({'success': True, 'id': record_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/records/<int:record_id>', methods=['PUT'])
@login_required
def api_update_record(record_id):
    """更新记录"""
    data = request.get_json()
    user = session['user']
    
    # 验证记录存在且有权限修改
    record = get_record_by_id(record_id)
    if not record:
        return jsonify({'error': '记录不存在'}), 404
    
    # 只有创建人或管理员可以修改
    if record['created_by'] != user['name'] and user['role'] != 'admin':
        return jsonify({'error': '无权修改此记录'}), 403
    
    try:
        update_record(record_id, data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/records/<int:record_id>', methods=['DELETE'])
@login_required
def api_delete_record(record_id):
    """删除记录"""
    user = session['user']
    
    # 验证记录存在且有权限删除
    record = get_record_by_id(record_id)
    if not record:
        return jsonify({'error': '记录不存在'}), 404
    
    # 只有创建人或管理员可以删除
    if record['created_by'] != user['name'] and user['role'] != 'admin':
        return jsonify({'error': '无权删除此记录'}), 403
    
    try:
        delete_record(record_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats/dashboard')
@login_required
def api_dashboard_stats():
    """获取仪表盘统计数据"""
    stats = get_dashboard_stats()
    return jsonify(stats)


@app.route('/api/stats/me')
@login_required
def api_my_stats():
    """获取我的统计数据"""
    user = session['user']
    stats = get_user_stats(user['name'])
    return jsonify(stats)


@app.route('/api/export')
@login_required
def api_export():
    """导出Excel"""
    user = session['user']
    
    # 只有统计员和管理员可以导出
    if user['role'] not in ['admin', 'statistician']:
        return jsonify({'error': '无权导出数据'}), 403
    
    # 获取时间范围参数
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    
    if start_date and end_date:
        records = get_filtered_records(start_date, end_date)
    else:
        records = get_records_by_user(user['name'], user['role'])
    
    filename = f'异常记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    export_to_excel(records, filename)
    
    return jsonify({'success': True, 'filename': filename})


@app.route('/api/download/<filename>')
@login_required
def api_download(filename):
    """下载文件"""
    user = session['user']
    
    if user['role'] not in ['admin', 'statistician']:
        return jsonify({'error': '无权下载'}), 403
    
    return send_from_directory('.', filename, as_attachment=True)


@app.route('/api/users', methods=['POST'])
@login_required
def api_create_user():
    """添加用户（仅管理员）"""
    user = session['user']
    if user['role'] != 'admin':
        return jsonify({'error': '无权添加用户'}), 403
    
    data = request.get_json()
    name = data.get('name')
    role = data.get('role')
    
    if not name or not role:
        return jsonify({'error': '缺少参数'}), 400
    
    if add_user(name, role):
        return jsonify({'success': True})
    else:
        return jsonify({'error': '用户已存在或添加失败'}), 400


def get_filtered_records(start_date=None, end_date=None):
    """根据日期范围筛选记录"""
    db_type = get_db_type()
    conn = get_db()
    
    if db_type == 'postgresql':
        from psycopg2.extras import RealDictCursor
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        if start_date and end_date:
            cursor.execute('''
                SELECT * FROM records 
                WHERE date >= %s AND date <= %s
                ORDER BY date DESC, id DESC
            ''', (start_date, end_date))
        else:
            cursor.execute('SELECT * FROM records ORDER BY date DESC, id DESC')
        
        records = [dict(row) for row in cursor.fetchall()]
        cursor.close()
    else:
        cursor = conn.cursor()
        
        if start_date and end_date:
            cursor.execute('''
                SELECT * FROM records 
                WHERE date >= ? AND date <= ?
                ORDER BY date DESC, id DESC
            ''', (start_date, end_date))
        else:
            cursor.execute('SELECT * FROM records ORDER BY date DESC, id DESC')
        
        records = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    return records


def add_user(name, role):
    """添加用户"""
    db_type = get_db_type()
    
    if db_type == 'postgresql':
        import psycopg2
        
        conn = psycopg2.connect(DATABASE_URL)
        cursor = conn.cursor()
        
        try:
            cursor.execute('INSERT INTO users (name, role) VALUES (%s, %s)', (name, role))
            conn.commit()
            cursor.close()
            conn.close()
            return True
        except psycopg2.IntegrityError:
            conn.rollback()
            cursor.close()
            conn.close()
            return False
    else:
        conn = sqlite3.connect(SQLITE_DB)
        cursor = conn.cursor()
        
        try:
            cursor.execute('INSERT INTO users (name, role) VALUES (?, ?)', (name, role))
            conn.commit()
            conn.close()
            return True
        except sqlite3.IntegrityError:
            conn.close()
            return False


# ==================== 启动 ====================

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=DEBUG)
